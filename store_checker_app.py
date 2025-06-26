import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import os
from datetime import datetime
import json

# â¬‡ï¸ êµ¬ê¸€ ë“œë¼ì´ë¸Œ ì—…ë¡œë“œìš© ì¶”ê°€
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from oauth2client.service_account import ServiceAccountCredentials

# íŒŒì¼ ì €ì¥ ê²½ë¡œ
if not os.path.exists("photos"):
    os.makedirs("photos")
if not os.path.exists("output"):
    os.makedirs("output")

EXCEL_PATH = "output/store_checker_data.xlsx"

# âœ… ë“œë¼ì´ë¸Œ ì—…ë¡œë“œ í•¨ìˆ˜
def upload_to_drive(local_path, file_name):
    scope = ['https://www.googleapis.com/auth/drive']

    # ğŸ‘‡ ì‹œí¬ë¦¿ì— ì €ì¥ëœ gdrive_credentials ë¶ˆëŸ¬ì˜¤ê¸°
    json_key = json.loads(st.secrets["gdrive_credentials"])
    credentials = ServiceAccountCredentials.from_json_keyfile_dict(json_key, scope)

    gauth = GoogleAuth()
    gauth.credentials = credentials
    drive = GoogleDrive(gauth)

    folder_id = '1rrrt-OmAYA08FMmyw7qO2HApOgyC24LY'  # â† ì‚¬ìš©ìì— ë”°ë¼ ë³€ê²½ í•„ìš”

    # ê¸°ì¡´ ë™ì¼ íŒŒì¼ ì‚­ì œ
    file_list = drive.ListFile({'q': f"title='{file_name}' and '{folder_id}' in parents and trashed=false"}).GetList()
    for f in file_list:
        f.Delete()

    # ìƒˆë¡œ ì—…ë¡œë“œ
    f = drive.CreateFile({'title': file_name, 'parents': [{'id': folder_id}]})
    f.SetContentFile(local_path)
    f.Upload()

# ë¸Œëœë“œ/ì¹´í…Œê³ ë¦¬ ì„ íƒ
brands = ["í…ŒíŒ”", "í•„ë¦½ìŠ¤", "ë½ì•¤ë½", "ë„ë£¨ì½”", "ê¸°íƒ€"]
categories = ["ì£¼ë°©", "ìƒí™œìš©í’ˆ", "ê°€ì „", "ì„¸ì œ", "ê¸°íƒ€"]

st.title("ë§¤ì¥ ì§„ì—´ ì‚¬ì§„ ì…ë ¥ ì•±")

# ì‚¬ì§„ ì…ë ¥
col1, col2, col3 = st.columns(3)
with col1:
    full_photo = st.file_uploader("ì „ì²´ ì§„ì—´ì‚¬ì§„", type=["jpg", "jpeg", "png"], key="full")
with col2:
    line1_photo = st.file_uploader("1ë²ˆì¤„ ì‚¬ì§„", type=["jpg", "jpeg", "png"], key="line1")
with col3:
    line2_photo = st.file_uploader("2ë²ˆì¤„ ì‚¬ì§„", type=["jpg", "jpeg", "png"], key="line2")

# ì •ë³´ ì…ë ¥
line1 = st.text_input("1ë²ˆ ì¤„ ì œí’ˆëª…")
line2 = st.text_input("2ë²ˆ ì¤„ ì œí’ˆëª…")
brand = st.selectbox("ë¸Œëœë“œ ì„ íƒ", brands)
category = st.selectbox("ì¹´í…Œê³ ë¦¬ ì„ íƒ", categories)

# ì €ì¥ ë²„íŠ¼
if st.button("ì œì¶œí•˜ê¸°"):
    if full_photo and line1_photo and line2_photo:
        now = datetime.now().strftime("%Y%m%d_%H%M%S")

        full_path = f"photos/full_{now}.jpg"
        line1_path = f"photos/line1_{now}.jpg"
        line2_path = f"photos/line2_{now}.jpg"
        with open(full_path, "wb") as f:
            f.write(full_photo.getbuffer())
        with open(line1_path, "wb") as f:
            f.write(line1_photo.getbuffer())
        with open(line2_path, "wb") as f:
            f.write(line2_photo.getbuffer())

        # ì—‘ì…€ ì—´ê¸°/ë§Œë“¤ê¸°
        if os.path.exists(EXCEL_PATH):
            wb = load_workbook(EXCEL_PATH)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["ì…ë ¥ì‹œê°„", "ì „ì²´ì‚¬ì§„", "ì¤„1ì‚¬ì§„", "ì¤„2ì‚¬ì§„", "ì¤„1 ì œí’ˆ", "ì¤„2 ì œí’ˆ", "ë¸Œëœë“œ", "ì¹´í…Œê³ ë¦¬"])

        row_num = ws.max_row + 1
        ws.cell(row=row_num, column=1, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row_num, column=5, value=line1)
        ws.cell(row=row_num, column=6, value=line2)
        ws.cell(row=row_num, column=7, value=brand)
        ws.cell(row=row_num, column=8, value=category)

        def insert_image(path, col_letter):
            img = XLImage(path)
            img.width, img.height = 150, 113
            ws.add_image(img, f"{col_letter}{row_num}")

        insert_image(full_path, "B")
        insert_image(line1_path, "C")
        insert_image(line2_path, "D")

        wb.save(EXCEL_PATH)

        try:
            upload_to_drive(EXCEL_PATH, "store_checker_data.xlsx")
            st.success("âœ… ì €ì¥ ë° ë“œë¼ì´ë¸Œ ì—…ë¡œë“œ ì™„ë£Œ!")
        except Exception as e:
            st.error(f"âš ï¸ ë“œë¼ì´ë¸Œ ì—…ë¡œë“œ ì‹¤íŒ¨: {e}")
    else:
        st.warning("âš ï¸ ëª¨ë“  ì‚¬ì§„ì„ ì—…ë¡œë“œí•´ì£¼ì„¸ìš”.")

